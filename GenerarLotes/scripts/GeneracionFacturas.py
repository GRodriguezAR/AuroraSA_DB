# -*- coding: utf-8 -*-
import csv
import random
from datetime import datetime, timedelta
import uuid
from collections import defaultdict
import hashlib
import openpyxl


def generar_facturas(NUM_FACTURAS):
    random.seed()
    PROBABILIDAD_CLIENTE_0 = 0.25
    SUCURSALES = ['SJ1', 'SJ2', 'RM1', 'RM2', 'LD1', 'MD1', 'MD2', 'CP1', 'CP2', 'HA1']
    TIPO_FACTURA = ['A', 'B', 'C']
    MEDIOS_PAGO = ['Tarjeta de credito', 'Efectivo', 'Billetera Electronica']
    
    # Precalcular valores constantes
    FECHA_INICIO = datetime(2023, 3, 1)
    FECHA_FIN = datetime(2025, 2, 28)
    DELTA_SEGUNDOS = (FECHA_FIN - FECHA_INICIO).total_seconds()
    DIGITOS = '0123456789'

    # Cargar datos una sola vez
    wb = openpyxl.load_workbook("../data/Informacion_complementaria.xlsx", data_only=True)
    
    # Cargar legajos y clientes
    LEGAJOS = [row[0] for row in wb["empleados"].iter_rows(min_row=2, values_only=True)]
    CLIENTES = [row[3] for row in wb["clientes"].iter_rows(min_row=2, values_only=True)]

    # Cargar productos
    def cargar_productos():
        with open('otros/prod.csv', 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=';')
            next(reader)  # Saltar encabezado
            return [(row[0], float(row[1])) for row in reader if len(row) >= 2]

    # Generación masiva de códigos únicos
    def generar_codigos_unicos(cantidad):
        generados = set()
        while len(generados) < cantidad:
            codigo = f"{random.randint(100,999)}-{random.randint(10,99)}-{random.randint(1000,9999)}"
            generados.add(codigo)
        return generados

    # Generar números de tarjeta válidos
    def generar_numero_tarjeta():
        tipo = random.choice(['Visa', 'Mastercard'])
        longitud = 15 if tipo == 'Visa' else 16
        prefijo = '4' if tipo == 'Visa' else '5' + random.choice(['1', '5'])
        
        numero = prefijo + ''.join(random.choices(DIGITOS, k=longitud - len(prefijo)))
        return '-'.join([numero[i:i+4] for i in range(0, len(numero), 4)])

    # Generar todos los códigos de factura al inicio
    codigos_facturas = generar_codigos_unicos(NUM_FACTURAS)
    
    # Cargar productos una sola vez
    PRODUCTOS = cargar_productos()
    
    # Generar todas las filas en memoria
    filas = []
    for codigo in codigos_facturas:
        tipo = random.choice(TIPO_FACTURA)
        sucursal = random.choice(SUCURSALES)
        cliente = '0' if random.random() < PROBABILIDAD_CLIENTE_0 else random.choice(CLIENTES)
        fecha = FECHA_INICIO + timedelta(seconds=random.random() * DELTA_SEGUNDOS)
        medio_pago = random.choice(MEDIOS_PAGO)
        empleado = random.choice(LEGAJOS)
        
        # Generar detalles de pago
        if medio_pago == 'Tarjeta de credito':
            id_pago = generar_numero_tarjeta()
        elif medio_pago == 'Billetera Electronica':
            id_pago = hashlib.sha256(uuid.uuid4().bytes).hexdigest()
        else:
            id_pago = ''
        
        # Generar items de factura
        num_items = random.randint(1, 5)
        for _ in range(num_items):
            producto, precio = random.choice(PRODUCTOS)
            filas.append([
                codigo,
                tipo,
                sucursal,
                cliente,
                producto,
                precio,
                random.randint(1, 10),
                fecha.strftime('%Y-%m-%d %H:%M:%S'),
                medio_pago,
                empleado,
                id_pago
            ])

    # Escribir facturas.csv
    with open('../data/facturas.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['CodigoFactura','TipoFactura','Sucursal','Cliente','Producto','Precio',
                        'Cantidad','FechaHora','MedioPago','Empleado','DetallesPago'])
        writer.writerows(filas)

    print(f"Facturas generadas")