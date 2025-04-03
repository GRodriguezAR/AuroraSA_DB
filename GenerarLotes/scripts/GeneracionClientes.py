# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
import random
from faker import Faker

def generar_clientes(CANT_CLIENTES):
    path = '../data/Informacion_complementaria.xlsx'
    fake = Faker('es_AR')
    random.seed()
    GENERO = ['Masculino','Femenino']
    TIPO = ['Normal', 'Miembro']

    # Si existe la hoja clientes, se borra y se crea una nueva
    wb = load_workbook(filename = path)
    if "clientes" in wb.sheetnames:
        hoja = wb["clientes"]
        hoja.delete_rows(1, hoja.max_row) 
    else:
        hoja = wb.create_sheet("clientes")

    #Encabezados
    hoja.append(('Nombre', 'Apellido', 'Genero', 'Dni', 'Tipo', 'Puntos'))
    for _ in range(CANT_CLIENTES):
        tip = random.choice(TIPO)
        pt = random.randint(0,100)*10
        #Si el tipo es Normal, los puntos son 0
        if tip == 'Normal':         
            pt = ''
        cli = ((fake.first_name(), fake.last_name(), random.choice(GENERO), random.randint(20000000,47000000), tip, pt))
        hoja.append(cli)

    wb.save(path)
    print('Clientes generados')