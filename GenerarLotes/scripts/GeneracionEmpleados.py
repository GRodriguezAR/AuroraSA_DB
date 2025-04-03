# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
import random
from datetime import datetime, timedelta
from faker import Faker

def generar_empleados(CANT_EMPLEADOS):
    path = "../data/Informacion_complementaria.xlsx"
    fake = Faker("es_AR")
    random.seed()
    
    SUCURSALES = ["SJ1", "SJ2", "RM1", "RM2", "LD1", "MD1", "MD2", "CP1", "CP2", "HA1"]
    CARGOS = ["Cajero", "Supervisor", "Gerente de sucursal"]
    TURNOS = ["TM", "TT", "JC"]
    GENERO = ["Masculino", "Femenino"]
    CARGOS_NO_GERENTE = CARGOS[:2]

    def generar_fecha():
        ini = datetime(2021, 3, 1)
        fin = datetime(2023, 2, 28)
        return ini + (fin - ini) * random.random()

    #Generar CUIL valido
    def generar_cuil(dni, genero):
        prefijo = "20" if genero == "Masculino" else "27"
        dni_str = str(dni).zfill(8)

        def calcular_dv(prefijo):
            coeficientes = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
            numero = prefijo + dni_str
            suma = sum(int(digit) * coeficientes[i] for i, digit in enumerate(numero))
            return (11 - (suma % 11)) % 11

        dv = calcular_dv(prefijo)
        if dv == 10:
            prefijo = "23"
            dv = calcular_dv(prefijo)
        
        return f"{prefijo}-{dni_str}-{dv}"

    def limpiar_telefono(numero):
        return "11" + numero.replace("+54", "").replace(" ", "")[-8:]

    def generar_empleado(cargo, sucursal):
        leg = random.randint(1000, 9999)
        nom = fake.first_name()
        ap = fake.last_name()
        gen = random.choice(GENERO)
        dni = random.randint(20000000, 47000000)
        tel = limpiar_telefono(fake.phone_number())
        
        return (
            leg,
            nom,
            ap,
            gen,
            generar_cuil(dni, gen),
            tel,
            fake.street_address() + ", Buenos Aires",
            generar_fecha().strftime("%Y-%m-%d %H:%M:%S"),
            f"{nom.replace(' ', '')}{ap.replace(' ', '')}{str(dni)[-3:]}@gmail.com",
            f"{ap.replace(' ', '')}{leg}@auroraSA.com.ar",
            cargo,
            sucursal,
            random.choice(TURNOS)
        )

    # Configurar el archivo Excel
    wb = load_workbook(filename=path)
    if "empleados" in wb.sheetnames:
        hoja = wb["empleados"]
        hoja.delete_rows(1, hoja.max_row)
    else:
        hoja = wb.create_sheet("empleados")
    
    hoja.append(("Legajo", "Nombre", "Apellido", "Genero", "Cuil", "Telefono", 
                "Direccion", "FechaAlta", "EmailPersonal", "EmailEmpresa", 
                "Cargo", "Sucursal", "Turno"))

    # Generar gerentes para cada sucursal
    for sucursal in SUCURSALES:
        empleado = generar_empleado(CARGOS[-1], sucursal)
        hoja.append(empleado)

    # Generar otros empleados
    for _ in range(CANT_EMPLEADOS - len(SUCURSALES)):
        cargo = random.choice(CARGOS_NO_GERENTE)
        sucursal = random.choice(SUCURSALES)
        empleado = generar_empleado(cargo, sucursal)
        hoja.append(empleado)

    wb.save(path)
    print("Empleados generados")